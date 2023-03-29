from tkinter import *
from tkinter import messagebox
import openpyxl

# Criando a interface gráfica
root = Tk()
root.title("ATIVIDADE CRUD")

# Criando as labels e os campos de entrada
label_nome = Label(root, text="Nome:")
entry_nome = Entry(root)

label_idade = Label(root, text="Idade:")
entry_idade = Entry(root)

label_curso = Label(root, text="Curso:")
entry_curso = Entry(root)

# Posicionando os elementos na interface
label_nome.grid(row=0, column=0, padx=5, pady=5)
entry_nome.grid(row=0, column=1, padx=5, pady=5)

label_idade.grid(row=1, column=0, padx=5, pady=5)
entry_idade.grid(row=1, column=1, padx=5, pady=5)

label_curso.grid(row=2, column=0, padx=5, pady=5)
entry_curso.grid(row=2, column=1, padx=5, pady=5)

# Função para salvar os dados do aluno no Excel
def salvar_aluno():
    # Abre o arquivo Excel e a planilha "alunos"
    wb = openpyxl.load_workbook('alunos.xlsx')
    ws = wb['alunos']

    # Verifica a última linha preenchida na planilha
    last_row = ws.max_row

    # Insere os dados do aluno na próxima linha vazia
    ws.cell(row=last_row+1, column=1).value = entry_nome.get()
    ws.cell(row=last_row+1, column=2).value = entry_idade.get()
    ws.cell(row=last_row+1, column=3).value = entry_curso.get()

    # Salva as alterações no arquivo Excel
    wb.save('alunos.xlsx')

    # Limpa os campos de entrada
    entry_nome.delete(0, END)
    entry_idade.delete(0, END)
    entry_curso.delete(0, END)

    # Exibe uma mensagem de sucesso
    messagebox.showinfo("Sucesso", "Aluno cadastrado com sucesso!")

# Botão para salvar o cadastro do aluno
btn_salvar = Button(root, text="Salvar", command=salvar_aluno)
btn_salvar.grid(row=3, column=1, padx=5, pady=5)

# Função para pesquisar um aluno pelo nome
def pesquisar_aluno():
    # Abre o arquivo Excel e a planilha "alunos"
    wb = openpyxl.load_workbook('alunos.xlsx')
    ws = wb['alunos']

    # Percorre todas as linhas da planilha em busca do nome do aluno
    for row in range(1, ws.max_row+1):
        if ws.cell(row=row, column=1).value == entry_nome.get():
            # Preenche os campos de entrada com os dados do aluno encontrado
            entry_idade.delete(0, END)
            entry_idade.insert(0, ws.cell(row=row, column=2).value)

            entry_curso.delete(0, END)
            entry_curso.insert(0, ws.cell(row=row, column=3).value)

            # Exibe uma mensagem de sucesso
            messagebox.showinfo("Sucesso", "Aluno encontrado!")
            return

    # Se o aluno não for encontrado, exibe uma mensagem de erro
    messagebox.showerror("Erro", "Aluno não encontrado.")

# Botão para pesquisar um aluno
btn_pesquisar = Button(root, text="Pesquisar", command=pesquisar_aluno)
btn_pesquisar.grid(row=3, column=0, padx=5, pady=5)

# Função para listar todos os alunos cadastrados
def listar_alunos():
    # Abre o arquivo Excel e a planilha "alunos"
    wb = openpyxl.load_workbook('alunos.xlsx')
    ws = wb['alunos']

    # Cria uma lista vazia para armazenar os alunos cadastrados
    alunos = []

    # Percorre todas as linhas da planilha e adiciona os alunos na lista
    for row in range(2, ws.max_row+1):
        aluno = {
            'nome': ws.cell(row=row, column=1).value,
            'idade': ws.cell(row=row, column=2).value,
            'curso': ws.cell(row=row, column=3).value
        }
        alunos.append(aluno)

    # Exibe uma janela com a lista de alunos cadastrados
    alunos_window = Toplevel(root)
    alunos_window.title("Alunos Cadastrados")
    alunos_window.geometry("500x500")

    scrollbar = Scrollbar(alunos_window)
    scrollbar.pack(side=RIGHT, fill=Y)

    listbox = Listbox(alunos_window, yscrollcommand=scrollbar.set)
    for aluno in alunos:
        listbox.insert(END, f"{aluno['nome']} - {aluno['idade']} anos - {aluno['curso']}")
    listbox.pack(side=LEFT, fill=BOTH)

    scrollbar.config(command=listbox.yview)

    # Fecha o arquivo Excel
    wb.close()

# Botão para listar os alunos cadastrados
btn_listar = Button(root, text="Listar Alunos", command=listar_alunos)
btn_listar.grid(row=4, column=0, padx=5, pady=5)



def excluir_aluno():
    # Abre o arquivo Excel e a planilha "alunos"
    wb = openpyxl.load_workbook('alunos.xlsx')
    ws = wb['alunos']

    # Percorre todas as linhas da planilha em busca do nome do aluno
    for row in range(1, ws.max_row+1):
        if ws.cell(row=row, column=1).value == entry_nome.get():
            # Remove a linha correspondente ao aluno encontrado
            ws.delete_rows(row)

            # Salva as alterações no arquivo Excel
            wb.save('alunos.xlsx')

            # Limpa os campos de entrada
            entry_nome.delete(0, END)
            entry_idade.delete(0, END)
            entry_curso.delete(0, END)

            # Exibe uma mensagem de sucesso
            messagebox.showinfo("Sucesso", "Aluno excluído com sucesso!")
            return

    # Se o aluno não for encontrado, exibe uma mensagem de erro
    messagebox.showerror("Erro", "Aluno não encontrado.")


btn_excluir = Button(root, text="Excluir", command=excluir_aluno)
btn_excluir.grid(row=4, column=1, padx=5, pady=5)


def editar_aluno():
    # Abre o arquivo Excel e a planilha "alunos"
    wb = openpyxl.load_workbook('alunos.xlsx')
    ws = wb['alunos']

    # Percorre todas as linhas da planilha em busca do nome do aluno
    for row in range(1, ws.max_row+1):
        if ws.cell(row=row, column=1).value == entry_nome.get():
            # Atualiza os dados do aluno encontrado
            ws.cell(row=row, column=2).value = entry_idade.get()
            ws.cell(row=row, column=3).value = entry_curso.get()

            # Salva as alterações no arquivo Excel
            wb.save('alunos.xlsx')

            # Limpa os campos de entrada
            entry_nome.delete(0, END)
            entry_idade.delete(0, END)
            entry_curso.delete(0, END)

            # Exibe uma mensagem de sucesso
            messagebox.showinfo("Sucesso", "Aluno editado com sucesso!")
            return

    # Se o aluno não for encontrado, exibe uma mensagem de erro
    messagebox.showerror("Erro", "Aluno não encontrado.")


btn_editar = Button(root, text="Editar", command=editar_aluno)
btn_editar.grid(row=4, column=2, padx=5, pady=5)



root.mainloop()





